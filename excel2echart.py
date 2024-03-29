import json
import os
import re
import sys
from openpyxl import load_workbook


wb = load_workbook(
    filename=sys.argv[1], read_only=True)
ws = wb[wb.sheetnames[0]]
# print(ws.max_column)

option = {
    'xAxis': {
        'type': 'time'
    },
    'yAxis': {
        'type': 'value',
        'scale': True
    },
    'tooltip': {
        'trigger': 'axis',
        'order': 'valueDesc',
    },
    'grid': {
        'left': '40',
        'right': '120' if ws.max_column > 7 else '40',
        'top': '40',
        'bottom': '40'
    },
    # 当数据列为 1 时，不显示
    'legend': {
        # 'type': 'scroll',
        # 当数据列小于 7 个时时，数据标题置顶，否则置右
        'orient': 'vertical' if ws.max_column > 7 else 'horizontal',
        'padding': 0,
        'right': 10,
        'top': 20,
        'bottom': 20,
        'selector': True,
        'data': []
    } if ws.max_column > 1 else {},
    'series': []
}
i = 0
for row in ws.rows:
    j = 0
    if i == 0:
        for cell in row:
            if j > 0:
                try:
                    option['legend']['data'].append(cell.value)
                except KeyError:
                    pass
                option['series'].append({
                    'name': cell.value,
                    'type': 'line',
                    'smooth': True,
                    'data': []
                })
            j += 1
    else:
        for cell in row:
            if j == 0:
                time_name = cell.value.strftime('%Y-%m-%d')
            else:
                if cell.value:
                    option['series'][j - 1]['data'] \
                        .append([time_name, cell.value])
                if cell.value is None and (i + 1) == ws.max_row:
                    option['series'][j - 1]['data'] \
                        .append([time_name, max([k[1] for k in option['series'][j - 1]['data']])])
            j += 1
    i += 1
wb.close()

json_data = json.dumps(option, ensure_ascii=False, indent=None)

html_template = open(f'{os.path.dirname(__file__)}/chart.html.template', 'rt')
html_template_str = html_template.read()
html_template.close()

html_str = re.sub(r'\{\{option\}\}', json_data, html_template_str)
print(html_str)
